#!/usr/bin/env python3
"""
좌표 매핑 시스템

어노테이션 이미지의 픽셀 좌표와 GPS 위치 데이터를 매핑하여
기물의 실제 지리적 위치와 이미지 내 위치를 연결합니다.
"""

import sys
import numpy as np
import pandas as pd
import cv2
import json
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import logging
from openpyxl import load_workbook

# 프로젝트 경로 추가
sys.path.append(str(Path(__file__).parent))
from config.paths import path_manager

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class CoordinateMappingSystem:
    """좌표 매핑 시스템"""
    
    def __init__(self):
        self.datasets_path = path_manager.datasets
        self.output_path = path_manager.processed_data
        self.figures_path = path_manager.figures
        
        # 데이터 구조
        self.gps_data = None
        self.annotation_image = None
        self.detected_objects = []
        self.coordinate_mappings = []
        
        logger.info("좌표 매핑 시스템 초기화 완료")
    
    def _parse_coordinate(self, coord_str: str) -> float:
        """도분초 형식의 좌표를 십진도 형식으로 변환"""
        try:
            # "36.5933983 N" 또는 "129 30.557773 E" 형식 파싱
            coord_str = coord_str.strip()
            direction = coord_str[-1]  # N, S, E, W
            coord_part = coord_str[:-1].strip()
            
            if ' ' in coord_part:
                # "129 30.557773" 형식 (도 분.분초)
                parts = coord_part.split()
                degrees = float(parts[0])
                minutes = float(parts[1])
                decimal_degrees = degrees + minutes / 60.0
            else:
                # "36.5933983" 형식 (이미 십진도)
                decimal_degrees = float(coord_part)
            
            # 남위/서경인 경우 음수로 변환
            if direction in ['S', 'W']:
                decimal_degrees = -decimal_degrees
                
            return decimal_degrees
            
        except Exception as e:
            logger.error(f"좌표 파싱 실패 '{coord_str}': {e}")
            return 0.0

    def load_gps_data(self) -> bool:
        """GPS 위치 데이터 로드"""
        gps_file = self.datasets_path / 'Location_MDGPS.xlsx'
        
        try:
            # 엑셀 파일 읽기
            workbook = load_workbook(gps_file)
            sheet = workbook.active
            
            data = []
            headers = [cell.value for cell in sheet[1]]
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):
                    data.append(list(row))
            
            self.gps_data = pd.DataFrame(data, columns=headers)
            
            # 데이터 정리
            self.gps_data = self.gps_data.dropna()
            
            # 좌표 형식 변환
            self.gps_data['위도_십진도'] = self.gps_data['위도'].apply(self._parse_coordinate)
            self.gps_data['경도_십진도'] = self.gps_data['경도'].apply(self._parse_coordinate)
            
            logger.info(f"GPS 데이터 로드 완료: {len(self.gps_data)}개 위치")
            logger.info(f"컬럼: {list(self.gps_data.columns)}")
            
            # 데이터 미리보기
            if len(self.gps_data) > 0:
                logger.info("GPS 데이터 샘플:")
                for i in range(min(5, len(self.gps_data))):
                    row = self.gps_data.iloc[i]
                    logger.info(f"  {row['정점']}: {row['위도_십진도']:.6f}, {row['경도_십진도']:.6f}")
            
            return True
            
        except Exception as e:
            logger.error(f"GPS 데이터 로드 실패: {e}")
            return False
    
    def load_detected_objects(self) -> bool:
        """이전에 감지된 객체 위치 정보 로드"""
        objects_file = self.figures_path / 'object_locations' / 'detected_objects.json'
        
        try:
            with open(objects_file, 'r') as f:
                self.detected_objects = json.load(f)
            
            logger.info(f"감지된 객체 로드 완료: {len(self.detected_objects)}개")
            return True
            
        except Exception as e:
            logger.error(f"객체 데이터 로드 실패: {e}")
            return False
    
    def load_annotation_image(self) -> bool:
        """어노테이션 이미지 로드"""
        annotation_file = self.datasets_path / 'PH_annotation.png'
        
        try:
            image = cv2.imread(str(annotation_file))
            if image is not None:
                self.annotation_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                logger.info(f"어노테이션 이미지 로드 완료: {self.annotation_image.shape}")
                return True
            else:
                logger.error("어노테이션 이미지 로드 실패")
                return False
                
        except Exception as e:
            logger.error(f"어노테이션 이미지 로드 실패: {e}")
            return False
    
    def create_coordinate_mapping(self) -> bool:
        """픽셀 좌표와 GPS 좌표 간의 매핑 생성"""
        if not self.gps_data is not None or not self.detected_objects:
            logger.error("GPS 데이터 또는 감지된 객체 데이터가 없습니다")
            return False
        
        try:
            # 이미지 크기 정보
            image_height, image_width = self.annotation_image.shape[:2]
            logger.info(f"이미지 크기: {image_width} x {image_height}")
            
            # GPS 데이터에서 위경도 범위 계산
            lat_min = self.gps_data['위도_십진도'].min()
            lat_max = self.gps_data['위도_십진도'].max()
            lon_min = self.gps_data['경도_십진도'].min()
            lon_max = self.gps_data['경도_십진도'].max()
            
            logger.info(f"GPS 좌표 범위:")
            logger.info(f"  위도: {lat_min:.6f} ~ {lat_max:.6f}")
            logger.info(f"  경도: {lon_min:.6f} ~ {lon_max:.6f}")
            
            # 객체들을 Y 좌표(깊이) 순으로 정렬 (이미지 상단부터 하단까지)
            sorted_objects = sorted(self.detected_objects, key=lambda obj: obj['center_y'])
            
            # GPS 데이터도 정점 번호 순으로 정렬 (1부터 25까지)
            gps_sorted = self.gps_data.sort_values('정점').reset_index(drop=True)
            
            logger.info(f"정렬된 GPS 데이터 순서:")
            for i, row in gps_sorted.iterrows():
                logger.info(f"  정점 {row['정점']}: ({row['위도_십진도']:.6f}, {row['경도_십진도']:.6f})")
            
            # 각 감지된 객체에 GPS 좌표 매핑
            mappings = []
            
            for i, obj in enumerate(sorted_objects):
                # GPS 데이터의 인덱스와 매핑 (1:1 매핑 시도)
                if i < len(gps_sorted):
                    gps_row = gps_sorted.iloc[i]
                    
                    mapping = {
                        'object_id': obj['id'],
                        'pixel_x': obj['center_x'],
                        'pixel_y': obj['center_y'],
                        'bbox': {
                            'x': obj['x'],
                            'y': obj['y'], 
                            'width': obj['width'],
                            'height': obj['height']
                        },
                        'gps_point_id': str(gps_row['정점']),
                        'latitude': float(gps_row['위도_십진도']),
                        'longitude': float(gps_row['경도_십진도']),
                        'mapping_confidence': self._calculate_mapping_confidence(obj, i, len(gps_sorted))
                    }
                    
                    mappings.append(mapping)
                    
                    logger.info(f"매핑 {i+1}: 객체 {obj['id']} -> GPS 정점 {gps_row['정점']} ({gps_row['위도_십진도']:.6f}, {gps_row['경도_십진도']:.6f})")
            
            self.coordinate_mappings = mappings
            logger.info(f"좌표 매핑 생성 완료: {len(mappings)}개")
            
            return True
            
        except Exception as e:
            logger.error(f"좌표 매핑 생성 실패: {e}")
            return False
    
    def _calculate_mapping_confidence(self, obj: Dict, index: int, total_gps: int) -> float:
        """매핑 신뢰도 계산"""
        try:
            # 간단한 신뢰도 계산 로직
            # 객체 크기 기반 신뢰도
            area_confidence = min(obj['area'] / 5000, 1.0)  # 큰 객체일수록 높은 신뢰도
            
            # 순서 기반 신뢰도 (처음과 끝은 높은 신뢰도)
            position_confidence = 1.0 if index < total_gps else 0.5
            
            # 전체 신뢰도
            confidence = (area_confidence + position_confidence) / 2.0
            
            return round(confidence, 3)
            
        except:
            return 0.5
    
    def save_coordinate_mappings(self) -> bool:
        """좌표 매핑 결과 저장"""
        try:
            # 출력 디렉토리 생성
            mapping_dir = self.output_path / 'coordinate_mappings'
            mapping_dir.mkdir(exist_ok=True)
            
            # JSON 형태로 저장
            mapping_file = mapping_dir / 'pixel_gps_mappings.json'
            with open(mapping_file, 'w') as f:
                json.dump(self.coordinate_mappings, f, indent=2, ensure_ascii=False)
            
            # CSV 형태로도 저장
            csv_file = mapping_dir / 'pixel_gps_mappings.csv'
            df_mappings = pd.DataFrame(self.coordinate_mappings)
            
            # bbox 정보를 별도 컬럼으로 분리
            df_mappings['bbox_x'] = df_mappings['bbox'].apply(lambda x: x['x'])
            df_mappings['bbox_y'] = df_mappings['bbox'].apply(lambda x: x['y'])
            df_mappings['bbox_width'] = df_mappings['bbox'].apply(lambda x: x['width'])
            df_mappings['bbox_height'] = df_mappings['bbox'].apply(lambda x: x['height'])
            df_mappings = df_mappings.drop('bbox', axis=1)
            
            df_mappings.to_csv(csv_file, index=False, encoding='utf-8-sig')
            
            logger.info(f"좌표 매핑 저장 완료: {mapping_dir}")
            return True
            
        except Exception as e:
            logger.error(f"좌표 매핑 저장 실패: {e}")
            return False
    
    def create_mapping_visualization(self) -> bool:
        """매핑 결과 시각화"""
        if not self.coordinate_mappings or self.annotation_image is None:
            logger.warning("매핑 데이터 또는 어노테이션 이미지가 없습니다")
            return False
        
        try:
            viz_dir = self.figures_path / 'coordinate_mappings'
            viz_dir.mkdir(exist_ok=True)
            
            fig, axes = plt.subplots(2, 2, figsize=(20, 16))
            
            # 1. 어노테이션 이미지 + 매핑된 객체들
            ax1 = axes[0, 0]
            ax1.imshow(self.annotation_image)
            ax1.set_title(f'Detected Objects with GPS Mapping ({len(self.coordinate_mappings)} objects)', fontsize=14)
            
            # 매핑된 객체들에 바운딩 박스와 GPS 정점 번호 표시
            for mapping in self.coordinate_mappings:
                bbox = mapping['bbox']
                rect = patches.Rectangle(
                    (bbox['x'], bbox['y']), 
                    bbox['width'], 
                    bbox['height'],
                    linewidth=2, 
                    edgecolor='yellow', 
                    facecolor='none'
                )
                ax1.add_patch(rect)
                
                # GPS 정점 번호 표시
                ax1.text(
                    mapping['pixel_x'], 
                    bbox['y'] - 5,
                    f"GPS-{mapping['gps_point_id']}", 
                    color='white',
                    fontsize=8,
                    ha='center',
                    fontweight='bold',
                    bbox=dict(boxstyle="round,pad=0.3", facecolor='black', alpha=0.7)
                )
            
            ax1.axis('off')
            
            # 2. GPS 좌표 분포
            ax2 = axes[0, 1]
            if len(self.coordinate_mappings) > 0:
                lats = [m['latitude'] for m in self.coordinate_mappings]
                lons = [m['longitude'] for m in self.coordinate_mappings]
                point_ids = [m['gps_point_id'] for m in self.coordinate_mappings]
                point_numbers = [int(pid.split('_')[1]) for pid in point_ids]  # PH_01 -> 1
                
                scatter = ax2.scatter(lons, lats, c=point_numbers, cmap='viridis', s=100, alpha=0.7)
                
                # 점 번호 표시
                for i, (lon, lat, point_id) in enumerate(zip(lons, lats, point_ids)):
                    ax2.annotate(f'{point_id}', (lon, lat), xytext=(5, 5), 
                               textcoords='offset points', fontsize=8)
                
                plt.colorbar(scatter, ax=ax2, label='GPS Point ID')
                ax2.set_xlabel('Longitude')
                ax2.set_ylabel('Latitude')
                ax2.set_title('GPS Coordinates Distribution', fontsize=14)
                ax2.grid(True, alpha=0.3)
            
            # 3. 픽셀 좌표 분포
            ax3 = axes[1, 0]
            if len(self.coordinate_mappings) > 0:
                pixel_xs = [m['pixel_x'] for m in self.coordinate_mappings]
                pixel_ys = [m['pixel_y'] for m in self.coordinate_mappings]
                confidences = [m['mapping_confidence'] for m in self.coordinate_mappings]
                
                scatter2 = ax3.scatter(pixel_xs, pixel_ys, c=confidences, cmap='RdYlGn', 
                                     s=100, alpha=0.7, vmin=0, vmax=1)
                
                plt.colorbar(scatter2, ax=ax3, label='Mapping Confidence')
                ax3.set_xlabel('Pixel X')
                ax3.set_ylabel('Pixel Y')
                ax3.set_title('Pixel Coordinates Distribution', fontsize=14)
                ax3.invert_yaxis()  # 이미지 좌표계에 맞춤
                ax3.grid(True, alpha=0.3)
            
            # 4. 매핑 신뢰도 히스토그램
            ax4 = axes[1, 1]
            if len(self.coordinate_mappings) > 0:
                confidences = [m['mapping_confidence'] for m in self.coordinate_mappings]
                ax4.hist(confidences, bins=10, alpha=0.7, color='skyblue', edgecolor='black')
                ax4.set_xlabel('Mapping Confidence')
                ax4.set_ylabel('Number of Objects')
                ax4.set_title('Mapping Confidence Distribution', fontsize=14)
                ax4.set_xlim(0, 1)
                
                # 통계 정보 표시
                mean_conf = np.mean(confidences)
                ax4.axvline(mean_conf, color='red', linestyle='--', 
                           label=f'Mean: {mean_conf:.3f}')
                ax4.legend()
            
            plt.tight_layout()
            plt.savefig(viz_dir / 'coordinate_mapping_analysis.png', dpi=150, bbox_inches='tight')
            plt.close()
            
            # 매핑 테이블 생성
            self._create_mapping_table(viz_dir)
            
            logger.info(f"매핑 시각화 생성 완료: {viz_dir}")
            return True
            
        except Exception as e:
            logger.error(f"매핑 시각화 생성 실패: {e}")
            return False
    
    def _create_mapping_table(self, output_dir: Path):
        """매핑 테이블 생성"""
        try:
            fig, ax = plt.subplots(figsize=(16, 10))
            ax.axis('tight')
            ax.axis('off')
            
            # 테이블 데이터 준비
            table_data = []
            headers = ['Object ID', 'Pixel (X, Y)', 'BBox Size', 'GPS Point', 'Lat, Lon', 'Confidence']
            
            for mapping in self.coordinate_mappings:
                row = [
                    mapping['object_id'],
                    f"({mapping['pixel_x']}, {mapping['pixel_y']})",
                    f"{mapping['bbox']['width']}×{mapping['bbox']['height']}",
                    mapping['gps_point_id'],
                    f"({mapping['latitude']:.6f}, {mapping['longitude']:.6f})",
                    f"{mapping['mapping_confidence']:.3f}"
                ]
                table_data.append(row)
            
            # 테이블 생성
            table = ax.table(cellText=table_data, colLabels=headers, 
                           cellLoc='center', loc='center')
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.2, 1.5)
            
            # 헤더 스타일링
            for i in range(len(headers)):
                table[(0, i)].set_facecolor('#40466e')
                table[(0, i)].set_text_props(weight='bold', color='white')
            
            # 신뢰도에 따른 행 색상
            for i, mapping in enumerate(self.coordinate_mappings):
                confidence = mapping['mapping_confidence']
                if confidence >= 0.8:
                    color = '#d4edda'  # 초록
                elif confidence >= 0.6:
                    color = '#fff3cd'  # 노랑
                else:
                    color = '#f8d7da'  # 빨강
                
                for j in range(len(headers)):
                    table[(i+1, j)].set_facecolor(color)
            
            plt.title('Pixel-GPS Coordinate Mapping Table', fontsize=16, fontweight='bold', pad=20)
            plt.savefig(output_dir / 'mapping_table.png', dpi=150, bbox_inches='tight')
            plt.close()
            
        except Exception as e:
            logger.error(f"매핑 테이블 생성 실패: {e}")
    
    def run_complete_mapping(self) -> bool:
        """전체 좌표 매핑 프로세스 실행"""
        logger.info("=== 좌표 매핑 시스템 실행 시작 ===")
        
        try:
            # 1. 데이터 로드
            logger.info("1. 필요 데이터 로드")
            if not self.load_gps_data():
                return False
                
            if not self.load_detected_objects():
                return False
                
            if not self.load_annotation_image():
                return False
            
            # 2. 좌표 매핑 생성
            logger.info("2. 좌표 매핑 생성")
            if not self.create_coordinate_mapping():
                return False
            
            # 3. 결과 저장
            logger.info("3. 매핑 결과 저장")
            if not self.save_coordinate_mappings():
                return False
            
            # 4. 시각화 생성
            logger.info("4. 매핑 결과 시각화")
            if not self.create_mapping_visualization():
                return False
            
            logger.info("=== 좌표 매핑 시스템 실행 완료 ===")
            self._print_mapping_summary()
            
            return True
            
        except Exception as e:
            logger.error(f"좌표 매핑 시스템 실행 실패: {e}")
            return False
    
    def _print_mapping_summary(self):
        """매핑 결과 요약 출력"""
        print("\n" + "="*60)
        print("좌표 매핑 결과 요약")
        print("="*60)
        
        print(f"📍 GPS 위치 데이터: {len(self.gps_data)}개")
        print(f"🎯 감지된 객체: {len(self.detected_objects)}개")
        print(f"🔗 생성된 매핑: {len(self.coordinate_mappings)}개")
        
        if self.coordinate_mappings:
            confidences = [m['mapping_confidence'] for m in self.coordinate_mappings]
            avg_confidence = np.mean(confidences)
            high_conf_count = sum(1 for c in confidences if c >= 0.8)
            
            print(f"📊 평균 신뢰도: {avg_confidence:.3f}")
            print(f"✅ 높은 신뢰도 매핑 (≥0.8): {high_conf_count}개")
        
        print(f"\n💾 출력 위치:")
        print(f"- 매핑 데이터: {self.output_path / 'coordinate_mappings'}")
        print(f"- 시각화: {self.figures_path / 'coordinate_mappings'}")
        
        print("="*60)


def main():
    """메인 실행 함수"""
    mapper = CoordinateMappingSystem()
    
    success = mapper.run_complete_mapping()
    
    if success:
        print("✅ 좌표 매핑 시스템이 성공적으로 완료되었습니다!")
        return 0
    else:
        print("❌ 좌표 매핑 시스템 실행 중 오류가 발생했습니다.")
        return 1


if __name__ == "__main__":
    exit(main())