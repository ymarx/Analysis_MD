#!/usr/bin/env python3
"""
향상된 실제 데이터 처리 파이프라인

datasets 폴더의 실제/모의 데이터에서 intensity data를 추출하고
기물 위치 매핑 및 특징 추출을 위한 전체 파이프라인을 구현합니다.
"""

import sys
import numpy as np
import pandas as pd
import cv2
from pathlib import Path
import json
import logging
from datetime import datetime
from PIL import Image
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from typing import List, Tuple, Dict, Optional, Union
import warnings
import pyxtf
from openpyxl import load_workbook

warnings.filterwarnings('ignore')

# 프로젝트 경로 추가
sys.path.append(str(Path(__file__).parent))

# 기존 모듈 임포트
from config.paths import path_manager

# 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EnhancedDataPipeline:
    """향상된 실제 데이터 처리 파이프라인"""
    
    def __init__(self):
        self.datasets_path = path_manager.datasets
        self.output_path = path_manager.processed_data
        self.figures_path = path_manager.figures
        
        # 데이터 구조체
        self.gps_data = None
        self.annotation_image = None
        self.object_locations = []
        self.intensity_data = {}
        
        logger.info("향상된 데이터 파이프라인 초기화 완료")

    def load_location_data(self) -> bool:
        """Location_MDGPS.xlsx 파일에서 기물 위치 정보 로드"""
        gps_file = self.datasets_path / 'Location_MDGPS.xlsx'
        
        try:
            # 엑셀 파일 로드 (openpyxl 사용)
            workbook = load_workbook(gps_file)
            sheet = workbook.active
            
            # 데이터를 리스트로 수집
            data = []
            headers = [cell.value for cell in sheet[1]]  # 첫 번째 행을 헤더로
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # 빈 행 제외
                    data.append(list(row))
            
            # DataFrame 생성
            self.gps_data = pd.DataFrame(data, columns=headers)
            
            logger.info(f"GPS 위치 데이터 로드 완료: {len(self.gps_data)}개 위치")
            logger.info(f"컬럼: {list(self.gps_data.columns)}")
            
            return True
            
        except Exception as e:
            logger.error(f"GPS 데이터 로드 실패: {e}")
            return False

    def load_annotation_image(self) -> bool:
        """PH_annotation.png 파일에서 어노테이션 이미지 로드"""
        annotation_file = self.datasets_path / 'PH_annotation.png'
        
        try:
            # OpenCV로 이미지 로드
            image = cv2.imread(str(annotation_file))
            if image is not None:
                self.annotation_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
                logger.info(f"어노테이션 이미지 로드 완료: {self.annotation_image.shape}")
                return True
            else:
                logger.error("어노테이션 이미지 로드 실패")
                return False
                
        except Exception as e:
            logger.error(f"어노테이션 이미지 로드 실패: {e}")
            return False

    def extract_object_locations_from_annotation(self) -> List[Dict]:
        """어노테이션 이미지에서 빨간 박스 위치 추출"""
        if self.annotation_image is None:
            logger.warning("어노테이션 이미지가 로드되지 않았습니다")
            return []
        
        try:
            # BGR로 변환 (OpenCV에서 사용하기 위해)
            image_bgr = cv2.cvtColor(self.annotation_image, cv2.COLOR_RGB2BGR)
            
            # HSV 변환하여 빨간색 영역 찾기
            hsv = cv2.cvtColor(image_bgr, cv2.COLOR_BGR2HSV)
            
            # 빨간색 범위 정의 (HSV)
            lower_red1 = np.array([0, 50, 50])
            upper_red1 = np.array([10, 255, 255])
            lower_red2 = np.array([170, 50, 50])
            upper_red2 = np.array([180, 255, 255])
            
            # 빨간색 마스크 생성
            mask1 = cv2.inRange(hsv, lower_red1, upper_red1)
            mask2 = cv2.inRange(hsv, lower_red2, upper_red2)
            mask = mask1 + mask2
            
            # 컨투어 찾기
            contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
            
            objects = []
            for i, contour in enumerate(contours):
                # 바운딩 박스 계산
                x, y, w, h = cv2.boundingRect(contour)
                
                # 너무 작은 영역 제외
                if w > 5 and h > 5:
                    objects.append({
                        'id': i + 1,
                        'x': int(x),
                        'y': int(y),
                        'width': int(w),
                        'height': int(h),
                        'center_x': int(x + w // 2),
                        'center_y': int(y + h // 2),
                        'area': int(w * h)
                    })
            
            # Y 좌표 기준으로 정렬 (위에서 아래로)
            objects.sort(key=lambda obj: obj['center_y'])
            
            # ID 재부여
            for i, obj in enumerate(objects):
                obj['id'] = i + 1
            
            self.object_locations = objects
            logger.info(f"어노테이션에서 {len(objects)}개의 객체 위치 추출 완료")
            
            return objects
            
        except Exception as e:
            logger.error(f"객체 위치 추출 실패: {e}")
            return []

    def list_available_datasets(self) -> Dict[str, Dict]:
        """사용 가능한 데이터셋 목록 조사"""
        datasets = {}
        
        for dataset_dir in self.datasets_path.iterdir():
            if dataset_dir.is_dir() and 'Pohang' in dataset_dir.name:
                dataset_info = {
                    'name': dataset_dir.name,
                    'path': dataset_dir,
                    'original_files': [],
                    'simulation_files': []
                }
                
                # Original 데이터 찾기
                original_path = dataset_dir / 'original'
                if original_path.exists():
                    for xtf_file in original_path.glob('*.xtf'):
                        dataset_info['original_files'].append(xtf_file)
                
                # Simulation 데이터 찾기
                simulation_path = dataset_dir / 'simulation'
                if simulation_path.exists():
                    for sim_dir in simulation_path.iterdir():
                        if sim_dir.is_dir():
                            for xtf_file in sim_dir.glob('*.xtf'):
                                dataset_info['simulation_files'].append(xtf_file)
                
                datasets[dataset_dir.name] = dataset_info
                
        logger.info(f"발견된 데이터셋: {len(datasets)}개")
        for name, info in datasets.items():
            logger.info(f"- {name}: Original {len(info['original_files'])}개, Simulation {len(info['simulation_files'])}개")
        
        return datasets

    def extract_intensity_from_xtf(self, xtf_path: Path, max_pings: int = 1000) -> Optional[Dict]:
        """XTF 파일에서 intensity 데이터 추출"""
        try:
            logger.info(f"XTF 파일 처리 시작: {xtf_path.name}")
            
            # pyxtf로 파일 읽기 (verbose 인자 제거)
            try:
                file_header, packets = pyxtf.xtf_read(str(xtf_path))
            except TypeError:
                # 이전 버전의 pyxtf에서 verbose 인자가 있는 경우
                file_header, packets = pyxtf.xtf_read(str(xtf_path), verbose=False)
            
            if not packets:
                logger.warning(f"패킷이 없는 파일: {xtf_path.name}")
                return None
            
            # 소나 패킷 필터링
            sonar_packets = [p for p in packets if hasattr(p, 'data')]
            
            if not sonar_packets:
                logger.warning(f"소나 데이터가 없는 파일: {xtf_path.name}")
                return None
            
            # 최대 ping 수 제한
            if len(sonar_packets) > max_pings:
                sonar_packets = sonar_packets[:max_pings]
                logger.info(f"패킷 수 제한: {max_pings}개로 제한됨")
            
            # intensity 매트릭스 구성
            intensity_data = []
            ping_info = []
            
            for i, packet in enumerate(sonar_packets):
                if hasattr(packet, 'data') and packet.data is not None:
                    # 데이터를 1차원 배열로 변환
                    if isinstance(packet.data, np.ndarray):
                        intensity_row = packet.data.flatten()
                    else:
                        intensity_row = np.array(packet.data).flatten()
                    
                    intensity_data.append(intensity_row)
                    
                    # ping 정보 수집
                    ping_info.append({
                        'ping_number': i,
                        'timestamp': getattr(packet, 'timestamp', None),
                        'latitude': getattr(packet, 'SensorYcoordinate', 0),
                        'longitude': getattr(packet, 'SensorXcoordinate', 0),
                        'samples': len(intensity_row)
                    })
            
            if not intensity_data:
                logger.warning(f"유효한 intensity 데이터가 없는 파일: {xtf_path.name}")
                return None
            
            # 모든 행의 길이를 통일 (최대 길이로 맞춤)
            max_samples = max(len(row) for row in intensity_data)
            normalized_data = []
            
            for row in intensity_data:
                if len(row) < max_samples:
                    # 부족한 부분을 0으로 패딩
                    padded_row = np.zeros(max_samples)
                    padded_row[:len(row)] = row
                    normalized_data.append(padded_row)
                else:
                    normalized_data.append(row[:max_samples])
            
            intensity_matrix = np.array(normalized_data)
            
            result = {
                'filename': xtf_path.name,
                'filepath': str(xtf_path),
                'intensity_matrix': intensity_matrix,
                'ping_info': ping_info,
                'shape': intensity_matrix.shape,
                'data_type': intensity_matrix.dtype,
                'processed_time': datetime.now()
            }
            
            logger.info(f"Intensity 데이터 추출 완료: {intensity_matrix.shape}")
            return result
            
        except Exception as e:
            logger.error(f"XTF 처리 실패 {xtf_path.name}: {e}")
            return None

    def process_all_datasets(self) -> bool:
        """모든 데이터셋에서 intensity 데이터 추출"""
        datasets = self.list_available_datasets()
        
        if not datasets:
            logger.error("처리할 데이터셋이 없습니다")
            return False
        
        all_results = {}
        
        for dataset_name, dataset_info in datasets.items():
            logger.info(f"데이터셋 처리 시작: {dataset_name}")
            
            dataset_results = {
                'name': dataset_name,
                'original_data': {},
                'simulation_data': {},
                'processing_time': datetime.now()
            }
            
            # Original 데이터 처리
            for xtf_file in dataset_info['original_files']:
                result = self.extract_intensity_from_xtf(xtf_file)
                if result:
                    key = xtf_file.name
                    dataset_results['original_data'][key] = result
            
            # Simulation 데이터 처리
            for xtf_file in dataset_info['simulation_files']:
                result = self.extract_intensity_from_xtf(xtf_file)
                if result:
                    key = xtf_file.name
                    dataset_results['simulation_data'][key] = result
            
            all_results[dataset_name] = dataset_results
            
            logger.info(f"데이터셋 {dataset_name} 처리 완료")
            logger.info(f"- Original: {len(dataset_results['original_data'])}개 파일")
            logger.info(f"- Simulation: {len(dataset_results['simulation_data'])}개 파일")
        
        self.intensity_data = all_results
        return True

    def save_intensity_data(self) -> bool:
        """추출된 intensity 데이터를 파일로 저장"""
        try:
            save_dir = self.output_path / 'intensity_data'
            save_dir.mkdir(exist_ok=True)
            
            for dataset_name, dataset_data in self.intensity_data.items():
                dataset_dir = save_dir / dataset_name
                dataset_dir.mkdir(exist_ok=True)
                
                # Original 데이터 저장
                original_dir = dataset_dir / 'original'
                original_dir.mkdir(exist_ok=True)
                
                for filename, data in dataset_data['original_data'].items():
                    save_path = original_dir / f"{Path(filename).stem}_intensity.npz"
                    np.savez_compressed(
                        save_path,
                        intensity_matrix=data['intensity_matrix'],
                        ping_info=data['ping_info'],
                        metadata={
                            'filename': data['filename'],
                            'shape': data['shape'],
                            'processed_time': str(data['processed_time'])
                        }
                    )
                
                # Simulation 데이터 저장
                simulation_dir = dataset_dir / 'simulation'
                simulation_dir.mkdir(exist_ok=True)
                
                for filename, data in dataset_data['simulation_data'].items():
                    save_path = simulation_dir / f"{Path(filename).stem}_intensity.npz"
                    np.savez_compressed(
                        save_path,
                        intensity_matrix=data['intensity_matrix'],
                        ping_info=data['ping_info'],
                        metadata={
                            'filename': data['filename'],
                            'shape': data['shape'],
                            'processed_time': str(data['processed_time'])
                        }
                    )
            
            logger.info(f"Intensity 데이터 저장 완료: {save_dir}")
            return True
            
        except Exception as e:
            logger.error(f"데이터 저장 실패: {e}")
            return False

    def generate_intensity_images(self, max_files_per_dataset: int = 2) -> bool:
        """Intensity 데이터를 이미지로 변환하여 저장"""
        try:
            images_dir = self.figures_path / 'intensity_images'
            images_dir.mkdir(exist_ok=True)
            
            for dataset_name, dataset_data in self.intensity_data.items():
                dataset_images_dir = images_dir / dataset_name
                dataset_images_dir.mkdir(exist_ok=True)
                
                # Original 데이터 이미지화
                original_files = list(dataset_data['original_data'].items())[:max_files_per_dataset]
                for filename, data in original_files:
                    self._create_intensity_image(
                        data['intensity_matrix'], 
                        dataset_images_dir / f"original_{Path(filename).stem}.png",
                        f"Original - {filename}"
                    )
                
                # Simulation 데이터 이미지화
                simulation_files = list(dataset_data['simulation_data'].items())[:max_files_per_dataset]
                for filename, data in simulation_files:
                    self._create_intensity_image(
                        data['intensity_matrix'], 
                        dataset_images_dir / f"simulation_{Path(filename).stem}.png",
                        f"Simulation - {filename}"
                    )
            
            logger.info(f"Intensity 이미지 생성 완료: {images_dir}")
            return True
            
        except Exception as e:
            logger.error(f"이미지 생성 실패: {e}")
            return False

    def _create_intensity_image(self, intensity_matrix: np.ndarray, save_path: Path, title: str):
        """Intensity 매트릭스를 이미지로 변환"""
        plt.figure(figsize=(12, 8))
        
        # 로그 스케일 적용
        log_intensity = np.log1p(np.abs(intensity_matrix))
        
        plt.imshow(log_intensity, cmap='hot', aspect='auto')
        plt.title(f'{title}\nShape: {intensity_matrix.shape}', fontsize=10)
        plt.xlabel('Sample Index')
        plt.ylabel('Ping Index')
        plt.colorbar(label='Log Intensity')
        
        plt.tight_layout()
        plt.savefig(save_path, dpi=150, bbox_inches='tight')
        plt.close()

    def create_object_location_overlay(self) -> bool:
        """어노테이션 이미지에 추출된 객체 위치를 오버레이"""
        if self.annotation_image is None or not self.object_locations:
            logger.warning("어노테이션 이미지 또는 객체 위치 정보가 없습니다")
            return False
        
        try:
            overlay_dir = self.figures_path / 'object_locations'
            overlay_dir.mkdir(exist_ok=True)
            
            # 원본 이미지 복사
            image_with_overlay = self.annotation_image.copy()
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(20, 12))
            
            # 원본 이미지
            ax1.imshow(self.annotation_image)
            ax1.set_title('Original Annotation Image', fontsize=14)
            ax1.axis('off')
            
            # 객체 위치 표시된 이미지
            ax2.imshow(image_with_overlay)
            ax2.set_title(f'Detected Objects: {len(self.object_locations)}', fontsize=14)
            
            # 감지된 객체에 바운딩 박스 그리기
            for obj in self.object_locations:
                rect = patches.Rectangle(
                    (obj['x'], obj['y']), 
                    obj['width'], 
                    obj['height'],
                    linewidth=2, 
                    edgecolor='lime', 
                    facecolor='none'
                )
                ax2.add_patch(rect)
                
                # ID 표시
                ax2.text(
                    obj['center_x'], 
                    obj['y'] - 5,
                    f"ID:{obj['id']}", 
                    color='yellow',
                    fontsize=10,
                    ha='center',
                    fontweight='bold'
                )
            
            ax2.axis('off')
            
            plt.tight_layout()
            plt.savefig(overlay_dir / 'object_detection_overlay.png', dpi=150, bbox_inches='tight')
            plt.close()
            
            # 객체 정보 JSON으로 저장
            with open(overlay_dir / 'detected_objects.json', 'w') as f:
                json.dump(self.object_locations, f, indent=2, default=str)
            
            logger.info(f"객체 위치 오버레이 생성 완료: {overlay_dir}")
            return True
            
        except Exception as e:
            logger.error(f"오버레이 생성 실패: {e}")
            return False

    def run_complete_pipeline(self) -> bool:
        """전체 파이프라인 실행"""
        logger.info("=== 향상된 데이터 파이프라인 실행 시작 ===")
        
        try:
            # 1. 기본 데이터 로드
            logger.info("1. 기본 데이터 로드")
            if not self.load_location_data():
                logger.error("GPS 데이터 로드 실패")
                return False
                
            if not self.load_annotation_image():
                logger.error("어노테이션 이미지 로드 실패")
                return False
            
            # 2. 객체 위치 추출
            logger.info("2. 어노테이션에서 객체 위치 추출")
            self.extract_object_locations_from_annotation()
            
            # 3. 모든 데이터셋에서 intensity 데이터 추출
            logger.info("3. 모든 데이터셋에서 intensity 데이터 추출")
            if not self.process_all_datasets():
                logger.error("데이터셋 처리 실패")
                return False
            
            # 4. 데이터 저장
            logger.info("4. 추출된 데이터 저장")
            if not self.save_intensity_data():
                logger.error("데이터 저장 실패")
                return False
            
            # 5. 이미지 생성
            logger.info("5. Intensity 이미지 생성")
            if not self.generate_intensity_images():
                logger.error("이미지 생성 실패")
                return False
            
            # 6. 객체 위치 오버레이 생성
            logger.info("6. 객체 위치 오버레이 생성")
            if not self.create_object_location_overlay():
                logger.error("오버레이 생성 실패")
                return False
            
            logger.info("=== 전체 파이프라인 실행 완료 ===")
            self._print_summary()
            
            return True
            
        except Exception as e:
            logger.error(f"파이프라인 실행 실패: {e}")
            return False

    def _print_summary(self):
        """실행 결과 요약 출력"""
        print("\n" + "="*60)
        print("데이터 파이프라인 실행 결과 요약")
        print("="*60)
        
        print(f"📍 GPS 위치 데이터: {len(self.gps_data) if self.gps_data is not None else 0}개 위치")
        print(f"🎯 감지된 객체: {len(self.object_locations)}개")
        print(f"📊 처리된 데이터셋: {len(self.intensity_data)}개")
        
        total_original = sum(len(data['original_data']) for data in self.intensity_data.values())
        total_simulation = sum(len(data['simulation_data']) for data in self.intensity_data.values())
        
        print(f"📁 Original 파일: {total_original}개")
        print(f"🔬 Simulation 파일: {total_simulation}개")
        
        print(f"\n💾 출력 위치:")
        print(f"- Intensity 데이터: {self.output_path / 'intensity_data'}")
        print(f"- 이미지: {self.figures_path / 'intensity_images'}")
        print(f"- 객체 위치: {self.figures_path / 'object_locations'}")
        
        print("="*60)


def main():
    """메인 실행 함수"""
    pipeline = EnhancedDataPipeline()
    
    success = pipeline.run_complete_pipeline()
    
    if success:
        print("✅ 데이터 파이프라인이 성공적으로 완료되었습니다!")
        return 0
    else:
        print("❌ 데이터 파이프라인 실행 중 오류가 발생했습니다.")
        return 1


if __name__ == "__main__":
    exit(main())